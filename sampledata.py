import openpyxl
from openpyxl import Workbook
import itertools

# wb = Workbook('sample_data.xlsx')
# wb.save('sample_data.xlsx')

wb = openpyxl.load_workbook('sample_data.xlsx')
ws = wb['Sheet']


def print_rows(ws):
    row_string = ""
    for row in ws.iter_rows(min_row=1, max_col=10, max_row=10):
        for cell in row:
            row_string += "{:<6}".format(str(cell.value) + " ")
            row_string += "\n"
            print(row_string)
